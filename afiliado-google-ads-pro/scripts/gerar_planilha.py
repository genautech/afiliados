#!/usr/bin/env python3
"""Gera a planilha de controle do afiliado (9 abas com fórmulas automáticas).

Uso:
  python3 gerar_planilha.py --comissao 97 --roas-meta 2.0 --moeda BRL \
      [--nicho "Idiomas"] -o planilha-afiliados.xlsx
"""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"; ACCENT = "E07B39"
HDR_FILL = PatternFill("solid", start_color=NAVY)
ALT_FILL = PatternFill("solid", start_color="EDF2F9")
YEL_FILL = PatternFill("solid", start_color="FFF2CC")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color=NAVY)
SUB_FONT = Font(name="Arial", italic=True, size=9, color="555555")
BODY = Font(name="Arial", size=10)
BLUE = Font(name="Arial", size=10, color="0000FF", bold=True)   # usuário preenche
CALC = Font(name="Arial", size=10, bold=True)                    # calcula sozinho
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def header(ws, row, cols):
    for c, txt in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=txt)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26


def title(ws, text, sub, ncols):
    ws["A1"] = text; ws["A1"].font = TITLE_FONT
    ws["A2"] = sub; ws["A2"].font = SUB_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.sheet_view.showGridLines = False


def widths(ws, w):
    for i, x in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


def body_rows(ws, start, n, ncols, blue_cols=(), wrap=()):
    for r in range(start, start + n):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN
            cell.font = BLUE if c in blue_cols else BODY
            cell.alignment = Alignment(vertical="top", wrap_text=(c in wrap))
            if (r - start) % 2 == 1:
                cell.fill = ALT_FILL


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--comissao", type=float, required=True)
    ap.add_argument("--roas-meta", type=float, default=2.0)
    ap.add_argument("--moeda", default="BRL", choices=["BRL", "USD", "GBP", "AUD"])
    ap.add_argument("--nicho", default="(defina na aba Nichos)")
    ap.add_argument("-o", "--out", default="planilha-afiliados.xlsx")
    a = ap.parse_args()
    simb = {"BRL": "R$", "USD": "US$", "GBP": "£", "AUD": "A$"}[a.moeda]
    numfmt = f'"{simb}" #,##0.00'

    wb = Workbook()

    # ===== 1. LEIA-ME =====
    ws = wb.active; ws.title = "Leia-me"
    title(ws, "PLANILHA DE CONTROLE — AFILIADOS COM GOOGLE ADS", f"Nicho: {a.nicho} · Moeda: {a.moeda} · Gerada pela skill afiliado-google-ads-pro", 3)
    widths(ws, [26, 100])
    ws["A4"] = "CONVENÇÕES"; ws["A4"].font = Font(name="Arial", bold=True, size=11, color=NAVY)
    conv = [("Células AZUIS", "Você preenche."),
            ("Células PRETAS (negrito)", "Calculam sozinhas — não editar."),
            ("Células AMARELAS", "Metas ajustáveis (comissão média, ROAS-meta)."),]
    for i, (k, v) in enumerate(conv, start=5):
        ws.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font = BODY
    ws["A9"] = "ORDEM DE USO"; ws["A9"].font = Font(name="Arial", bold=True, size=11, color=NAVY)
    ordem = [
        ("2. Nichos", "Escolha UM nicho pelo score e pela adequação ao Google Ads."),
        ("3. Produtos e Programas", "Selecione 1–2 ofertas. Valide comissões DENTRO da plataforma — mudam com frequência."),
        ("4. Palavras-chave", "Monte o banco e valide os CPCs reais no Planejador antes de definir lances."),
        ("5. Negativas", "Suba a lista-mestra como lista compartilhada ANTES do primeiro real gasto."),
        ("6. Estrutura de Campanhas", "Copie a nomenclatura e registre a fase de cada campanha."),
        ("7. Copy de Anúncios", "Redija com o contador automático — ele avisa se estourar o limite."),
        ("8. Controle de Orçamento", "Preencha DIARIAMENTE — CPA e ROAS calculam sozinhos."),
        ("9. Métricas Semanais", "Toda segunda: registre a semana e siga o status automático (✔/⚙/✖)."),
    ]
    for i, (k, v) in enumerate(ordem, start=10):
        ws.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font = BODY

    # ===== 2. NICHOS =====
    ws = wb.create_sheet("Nichos")
    title(ws, "AVALIAÇÃO DE NICHOS", "Pontue 0–10 cada critério; o score calcula sozinho. Escolha UM nicho ≥ 80 (ou 65+ com ângulo).", 10)
    widths(ws, [30, 12, 12, 12, 12, 12, 12, 10, 22, 30])
    header(ws, 4, ["Nicho", "Intenção (0-10)", "Economia (0-10)", "Adequação Ads (0-10)", "Durabilidade (0-10)", "Concorrência (0-10)", "Faceless/IA (0-10)", "SCORE", "Veredicto", "Notas"])
    for r in range(5, 15):
        ws.cell(row=r, column=8, value=f"=IF(B{r}=\"\",\"\",ROUND((B{r}*0.25+C{r}*0.25+D{r}*0.2+E{r}*0.15+F{r}*0.1+G{r}*0.05)*10,0))").font = CALC
        ws.cell(row=r, column=9, value=f"=IF(H{r}=\"\",\"\",IF(H{r}>=80,\"✔ PRIORIDADE\",IF(H{r}>=65,\"⚙ VIÁVEL C/ ÂNGULO\",\"✖ EVITAR\")))").font = CALC
    body_rows(ws, 5, 10, 10, blue_cols=(1, 2, 3, 4, 5, 6, 7), wrap=(1, 10))

    # ===== 3. PRODUTOS E PROGRAMAS =====
    ws = wb.create_sheet("Produtos e Programas")
    title(ws, "OFERTAS CANDIDATAS", "Comissão ≥ 3× o CPA estimado. Valide as taxas dentro da plataforma.", 9)
    widths(ws, [28, 16, 14, 14, 16, 16, 14, 14, 30])
    header(ws, 4, ["Produto/Oferta", "Plataforma", f"Comissão ({simb})", "Modelo (CPA/%/rec.)", f"CPA estimado ({simb})", "Passa na regra 3×?", "País aceito", "Status", "Notas/regras do produtor"])
    for r in range(5, 15):
        ws.cell(row=r, column=6, value=f"=IF(OR(C{r}=\"\",E{r}=\"\"),\"\",IF(C{r}>=3*E{r},\"✔ SIM\",\"✖ NÃO — não fecha a conta\"))").font = CALC
    body_rows(ws, 5, 10, 9, blue_cols=(1, 2, 3, 4, 5, 7, 8, 9), wrap=(1, 9))

    # ===== 4. PALAVRAS-CHAVE =====
    ws = wb.create_sheet("Palavras-chave")
    title(ws, "BANCO DE PALAVRAS-CHAVE", "CPCs estimados são planejamento — valide no Planejador (coluna F) antes do lance.", 8)
    widths(ws, [34, 16, 18, 16, 16, 16, 16, 26])
    header(ws, 4, ["Palavra-chave", "Grupo", "Intenção (transacional/comparativa/info)", "Correspondência", f"CPC estimado ({simb})", f"CPC real Planejador ({simb})", "Volume/mês", "Observações"])
    body_rows(ws, 5, 30, 8, blue_cols=(1, 2, 3, 4, 5, 6, 7, 8), wrap=(1, 8))

    # ===== 5. NEGATIVAS =====
    ws = wb.create_sheet("Negativas")
    title(ws, "LISTA-MESTRA DE NEGATIVAS", "Subir como lista compartilhada ANTES de ativar. Alimente com os termos de pesquisa a cada 3 dias.", 3)
    widths(ws, [30, 20, 40])
    header(ws, 4, ["Negativa", "Correspondência", "Origem (padrão / termos de pesquisa)"])
    negs = ["grátis", "gratuito", "free", "download", "baixar", "pdf", "torrent", "crackeado",
            "login", "entrar", "reclame aqui", "reclamação", "golpe", "fraude", "é confiável",
            "vagas", "emprego", "curso grátis", "telegram", "grupo whatsapp", "como fazer em casa",
            "receita caseira", "o que é", "significado", "wikipedia", "youtube"]
    for i, n in enumerate(negs, start=5):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value="frase")
        ws.cell(row=i, column=3, value="padrão da skill")
    body_rows(ws, 5, len(negs) + 10, 3, blue_cols=(1, 2, 3))

    # ===== 6. ESTRUTURA DE CAMPANHAS =====
    ws = wb.create_sheet("Estrutura de Campanhas")
    title(ws, "ESTRUTURA E FASES", "Nomenclatura: [TIPO]-[NICHO]-[OBJETIVO]-[FASE] · país prefixado fora do BR (US-PESQ-...)", 8)
    widths(ws, [34, 28, 30, 14, 16, 16, 14, 30])
    header(ws, 4, ["Campanha", "Grupo de anúncios", "Keyword principal", "Fase (1/2/3)", f"Orçamento/dia ({simb})", "Lance", "Status", "Próxima decisão + data"])
    body_rows(ws, 5, 20, 8, blue_cols=(1, 2, 3, 4, 5, 6, 7, 8), wrap=(1, 2, 3, 8))

    # ===== 7. COPY DE ANÚNCIOS =====
    ws = wb.create_sheet("Copy de Anúncios")
    title(ws, "BANCO DE COPY (RSA)", "Limites: título 30c · descrição 90c · caminho 15c. O contador avisa se estourar.", 5)
    widths(ws, [14, 70, 10, 12, 26])
    header(ws, 4, ["Tipo", "Texto", "Carac.", "Situação", "Ângulo/nota"])
    tipos = ["Título"] * 15 + ["Descrição"] * 4 + ["Caminho"] * 2
    for i, t in enumerate(tipos, start=5):
        ws.cell(row=i, column=1, value=t).font = BODY
        lim = 30 if t == "Título" else (90 if t == "Descrição" else 15)
        ws.cell(row=i, column=3, value=f"=IF(B{i}=\"\",\"\",LEN(B{i}))").font = CALC
        ws.cell(row=i, column=4, value=f"=IF(B{i}=\"\",\"\",IF(LEN(B{i})>{lim},\"✖ ESTOUROU ({lim}c)\",\"✔ OK\"))").font = CALC
    body_rows(ws, 5, len(tipos), 5, blue_cols=(2, 5), wrap=(2, 5))

    # ===== 8. CONTROLE DE ORÇAMENTO (diário) =====
    ws = wb.create_sheet("Controle de Orçamento")
    title(ws, "CONTROLE DIÁRIO", "Preencha as colunas azuis todo dia. O resto calcula sozinho a partir da comissão média (célula C4).", 11)
    widths(ws, [14, 30, 18, 14, 14, 14, 12, 12, 14, 14, 12])
    ws["B4"] = "Comissão média por venda:"; ws["B4"].font = Font(name="Arial", bold=True, size=10)
    c = ws["C4"]; c.value = a.comissao; c.fill = YEL_FILL; c.font = Font(name="Arial", bold=True, size=10); c.number_format = numfmt; c.border = THIN
    ws["D4"] = "ROAS-meta:"; ws["D4"].font = Font(name="Arial", bold=True, size=10)
    e = ws["E4"]; e.value = a.roas_meta; e.fill = YEL_FILL; e.font = Font(name="Arial", bold=True, size=10); e.border = THIN
    ws["F4"] = "CPA-alvo:"; ws["F4"].font = Font(name="Arial", bold=True, size=10)
    g = ws["G4"]; g.value = "=IF(E4=0,\"\",C4/E4)"; g.font = CALC; g.number_format = numfmt; g.border = THIN
    header(ws, 6, ["Data", "Campanha", f"Gasto ({simb})", "Impressões", "Cliques", "Conversões", "CTR", f"CPC ({simb})", f"CPA ({simb})", f"Receita ({simb})", "ROAS"])
    for r in range(7, 97):
        ws.cell(row=r, column=7, value=f"=IF(OR(D{r}=\"\",D{r}=0),\"\",E{r}/D{r})").font = CALC
        ws.cell(row=r, column=7).number_format = "0.00%"
        ws.cell(row=r, column=8, value=f"=IF(OR(E{r}=\"\",E{r}=0),\"\",C{r}/E{r})").font = CALC
        ws.cell(row=r, column=8).number_format = numfmt
        ws.cell(row=r, column=9, value=f"=IF(OR(F{r}=\"\",F{r}=0),\"—\",C{r}/F{r})").font = CALC
        ws.cell(row=r, column=9).number_format = numfmt
        ws.cell(row=r, column=10, value=f"=IF(F{r}=\"\",\"\",F{r}*$C$4)").font = CALC
        ws.cell(row=r, column=10).number_format = numfmt
        ws.cell(row=r, column=11, value=f"=IF(OR(C{r}=\"\",C{r}=0),\"\",J{r}/C{r})").font = CALC
        ws.cell(row=r, column=11).number_format = "0.00"
        for col in (3,):
            ws.cell(row=r, column=col).number_format = numfmt
    body_rows(ws, 7, 90, 11, blue_cols=(1, 2, 3, 4, 5, 6))

    # ===== 9. MÉTRICAS SEMANAIS =====
    ws = wb.create_sheet("Métricas Semanais")
    title(ws, "PAINEL SEMANAL DE KPIs", "Toda segunda-feira. O status aplica as regras: ROAS ≥ meta = ✔ ESCALAR · 1.0–meta = ⚙ OTIMIZAR · 3× comissão sem venda = ✖ PAUSAR/TROCAR.", 10)
    widths(ws, [16, 28, 16, 14, 14, 14, 14, 12, 20, 30])
    header(ws, 4, ["Semana (início)", "Campanha", f"Gasto ({simb})", "Cliques", "Conversões", f"CPA ({simb})", f"Receita ({simb})", "ROAS", "STATUS", "Ação da semana (UMA mudança)"])
    for r in range(5, 45):
        ws.cell(row=r, column=6, value=f"=IF(OR(E{r}=\"\",E{r}=0),\"—\",C{r}/E{r})").font = CALC
        ws.cell(row=r, column=6).number_format = numfmt
        ws.cell(row=r, column=7, value=f"=IF(E{r}=\"\",\"\",E{r}*'Controle de Orçamento'!$C$4)").font = CALC
        ws.cell(row=r, column=7).number_format = numfmt
        ws.cell(row=r, column=8, value=f"=IF(OR(C{r}=\"\",C{r}=0),\"\",G{r}/C{r})").font = CALC
        ws.cell(row=r, column=8).number_format = "0.00"
        ws.cell(row=r, column=9, value=(
            f"=IF(C{r}=\"\",\"\","
            f"IF(AND(OR(E{r}=0,E{r}=\"\"),C{r}>=3*'Controle de Orçamento'!$C$4),\"✖ PAUSAR/TROCAR OFERTA\","
            f"IF(E{r}=\"\",\"⏳ AGUARDAR AMOSTRA\","
            f"IF(H{r}>='Controle de Orçamento'!$E$4,\"✔ ESCALAR (+20-30%)\","
            f"IF(H{r}>=1,\"⚙ OTIMIZAR\",\"✖ CORTAR DESPERDÍCIO\")))))")).font = CALC
        ws.cell(row=r, column=3).number_format = numfmt
    body_rows(ws, 5, 40, 10, blue_cols=(1, 2, 3, 4, 5, 10), wrap=(2, 10))

    wb.save(a.out)
    print(f"Planilha gerada: {a.out} (9 abas · comissão {simb} {a.comissao} · ROAS-meta {a.roas_meta})")


if __name__ == "__main__":
    main()
